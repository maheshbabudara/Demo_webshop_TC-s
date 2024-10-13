import openpyxl
import openpyxl as p
from time import sleep

def reader(excel_name,sheet_name):
    workbook=p.load_workbook(excel_name)
    sheet=workbook[sheet_name]
    credentials=[]
    for row in range(1, sheet.max_row+1):
        cred=[]
        for col in range(1, sheet.max_column+1):
            data=sheet.cell(row,col)
            cred.append(data.value)
        credentials.append(cred)
    return



# def write(excel_name,sheet_name):
#     book=openpyxl.load_workbook(excel_name)
#     sheet=book[sheet_name]
#     sheet.cell(10,11).value='rony'
#     book.save(excel_name)

# write('file','sheetname')



